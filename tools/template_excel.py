#template_excel.py
"""
Génère le template Excel standard pour la saisie des données financières.
Ce template est proposé au téléchargement depuis l'interface graphique.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette couleurs ──────────────────────────────────────────────────────────
C_HEADER_BG   = "FF1F3864"
C_HEADER_FG   = "FFFFFFFF"
C_SECTION_BG  = "FFD6E4F0"
C_OPTIONAL_BG = "FFFFF2CC"
C_CALC_BG     = "FFE2EFDA"
C_BORDER      = "FF8EA9C1"

THIN  = Side(style="thin",   color=C_BORDER)
THICK = Side(style="medium", color="FF1F3864")

YEARS_HIST = ["N-3", "N-2", "N-1"]
YEARS_PROJ = ["N+1", "N+2", "N+3", "N+4", "N+5"]
ALL_YEARS  = YEARS_HIST + YEARS_PROJ

FONT_TITLE   = Font(name="Arial", bold=True, size=11, color=C_HEADER_FG)
FONT_SECTION = Font(name="Arial", bold=True, size=10, color="FF1F3864")
FONT_LABEL   = Font(name="Arial", size=9)
FONT_INPUT   = Font(name="Arial", size=9, color="FF0000FF")
FONT_FORMULA = Font(name="Arial", size=9, color="FF000000")

NUM_FMT_MAD  = '#,##0;(#,##0);"-"'
NUM_FMT_PCT  = '0.0%;-0.0%;"-"'
NUM_FMT_TEXT = '@'


def _border_all():
    return Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

def _border_left_thick():
    return Border(top=THIN, bottom=THIN, left=THICK, right=THIN)

def _header_fill():
    return PatternFill("solid", start_color=C_HEADER_BG)

def _section_fill():
    return PatternFill("solid", start_color=C_SECTION_BG)

def _optional_fill():
    return PatternFill("solid", start_color=C_OPTIONAL_BG)

def _calc_fill():
    return PatternFill("solid", start_color=C_CALC_BG)


def _build_financial_sheet(wb):
    ws = wb.active
    ws.title = "Données Financières"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 6
    for col in range(3, 11):
        ws.column_dimensions[get_column_letter(col)].width = 13

    COL_START = 3  # Colonne C

    # ── Titre principal ───────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "TEMPLATE D'ÉVALUATION FINANCIÈRE — MÉTHODES DCF & MULTIPLES"
    c.font = Font(name="Arial", bold=True, size=13, color=C_HEADER_FG)
    c.fill = _header_fill()
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── En-tête colonnes ─────────────────────────────────────────────────────
    ws.merge_cells("A2:B2")
    ws["A2"].value = "Indicateur"
    ws["A2"].font = FONT_TITLE
    ws["A2"].fill = _header_fill()
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    for i, yr in enumerate(ALL_YEARS):
        col = COL_START + i
        cell = ws.cell(row=2, column=col, value=yr)
        cell.font = FONT_TITLE
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border_left_thick() if yr == "N+1" else _border_all()

    ws.row_dimensions[2].height = 20

    # ── Légende historique / projection ──────────────────────────────────────
    ws.merge_cells("C3:E3")
    ws["C3"].value = "◀  Historique (à saisir)  ▶"
    ws["C3"].font = Font(name="Arial", size=8, italic=True, color="FF1F3864")
    ws["C3"].alignment = Alignment(horizontal="center")

    ws.merge_cells("F3:J3")
    ws["F3"].value = "◀  Projections calculées par l'agent  ▶"
    ws["F3"].font = Font(name="Arial", size=8, italic=True, color="FF375623")
    ws["F3"].alignment = Alignment(horizontal="center")

    row = [4]  # liste pour mutation dans les closures

    def section(title):
        ws.merge_cells(f"A{row[0]}:J{row[0]}")
        c = ws.cell(row=row[0], column=1, value=f"  {title}")
        c.font = FONT_SECTION
        c.fill = _section_fill()
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _border_all()
        ws.row_dimensions[row[0]].height = 18
        row[0] += 1

    def data_row(label, unit="kMAD", fmt=NUM_FMT_MAD, optional=False):
        r = row[0]

        # Label
        lbl = ws.cell(row=r, column=1, value=f"  {label}")
        lbl.font = FONT_LABEL
        lbl.border = _border_all()
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        if optional:
            lbl.fill = _optional_fill()

        # Unité
        u = ws.cell(row=r, column=2, value=unit)
        u.font = Font(name="Arial", size=8, italic=True, color="FF808080")
        u.alignment = Alignment(horizontal="center", vertical="center")
        u.border = _border_all()

        # Colonnes données
        for i, yr in enumerate(ALL_YEARS):
            col = COL_START + i
            cell = ws.cell(row=r, column=col)
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = _border_left_thick() if yr == "N+1" else _border_all()

            if yr.startswith("N+"):
                cell.fill = _calc_fill()
                cell.font = Font(name="Arial", size=8, italic=True, color="FF808080")
                cell.value = "→ Agent"
            else:
                cell.font = FONT_INPUT
                if optional:
                    cell.fill = _optional_fill()

        ws.row_dimensions[r].height = 17
        row[0] += 1

    # ── SECTION 1 — COMPTE DE RÉSULTAT ───────────────────────────────────────
    section("1. COMPTE DE RÉSULTAT (kMAD)")
    data_row("Chiffre d'affaires (CA)")
    data_row("Coût des marchandises vendues (CMV)")
    data_row("Marge brute")
    data_row("Charges d'exploitation (hors D&A)")
    data_row("Amortissements & dépréciations (D&A)")
    data_row("EBIT (Résultat d'exploitation)")
    data_row("Charges financières (intérêts)")
    data_row("Résultat avant impôt (EBT)")
    data_row("Impôt sur les sociétés")
    data_row("Résultat net")
    row[0] += 1

    # ── SECTION 2 — BILAN ────────────────────────────────────────────────────
    section("2. BILAN — Postes clés (kMAD)")
    data_row("Capitaux propres")
    data_row("Dettes financières à long terme")
    data_row("Dettes financières à court terme")
    data_row("Dette financière nette")
    data_row("Trésorerie & équivalents")
    data_row("Besoin en fonds de roulement (BFR)")
    data_row("Variation du BFR (ΔBFR)")
    row[0] += 1

    # ── SECTION 3 — FLUX / DCF ───────────────────────────────────────────────
    section("3. FLUX DE TRÉSORERIE — Données DCF (kMAD)")
    data_row("CAPEX (investissements)")
    data_row("EBITDA", optional=True)
    row[0] += 1

    # ── SECTION 4 — PARAMÈTRES WACC ──────────────────────────────────────────
    section("4. PARAMÈTRES WACC (optionnels — estimés par l'agent si vides)")
    data_row("Taux d'imposition effectif", unit="%", fmt=NUM_FMT_PCT, optional=True)
    data_row("Coût de la dette (%)",       unit="%", fmt=NUM_FMT_PCT, optional=True)
    data_row("Bêta (β)",                   unit="—", fmt='0.00',      optional=True)
    data_row("Prime de risque marché (%)", unit="%", fmt=NUM_FMT_PCT, optional=True)
    data_row("Taux sans risque (%)",       unit="%", fmt=NUM_FMT_PCT, optional=True)
    row[0] += 1

    # ── SECTION 5 — TAUX DE CROISSANCE ───────────────────────────────────────
    section("5. TAUX DE CROISSANCE (optionnel — estimé par l'agent si vide)")
    data_row("Taux de croissance court terme (N+1 à N+5)", unit="%",
             fmt=NUM_FMT_PCT, optional=True)
    data_row("Taux de croissance terminal (après N+5)",    unit="%",
             fmt=NUM_FMT_PCT, optional=True)
    row[0] += 1

    # ── SECTION 6 — INFORMATIONS GÉNÉRALES ───────────────────────────────────
    section("6. INFORMATIONS GÉNÉRALES")

    gen_items = [
        ("Nom de l'entreprise",         "Texte", NUM_FMT_TEXT),
        ("Secteur d'activité",          "Texte", NUM_FMT_TEXT),
        ("Nombre d'actions (milliers)", "k",     NUM_FMT_MAD),
        ("Prix de l'action (MAD)",      "MAD",   '#,##0.00'),
    ]
    for label, unit, fmt in gen_items:
        r = row[0]
        lbl = ws.cell(row=r, column=1, value=f"  {label}")
        lbl.font = FONT_LABEL
        lbl.border = _border_all()
        lbl.alignment = Alignment(horizontal="left", vertical="center")

        u = ws.cell(row=r, column=2, value=unit)
        u.font = Font(name="Arial", size=8, italic=True, color="FF808080")
        u.alignment = Alignment(horizontal="center", vertical="center")
        u.border = _border_all()

        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        inp = ws.cell(row=r, column=3)
        inp.font = FONT_INPUT
        inp.number_format = fmt
        inp.alignment = Alignment(horizontal="left", vertical="center")
        inp.border = _border_all()

        for col in range(6, 11):
            c = ws.cell(row=r, column=col)
            c.fill = _calc_fill()
            c.border = _border_left_thick() if col == 6 else _border_all()

        ws.row_dimensions[r].height = 17
        row[0] += 1

    row[0] += 1

    # ── Légende ───────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row[0]}:J{row[0]}")
    leg = ws[f"A{row[0]}"]
    leg.value = (
        "🔵 Bleu = Saisie obligatoire   "
        "🟡 Jaune = Optionnel (estimé par l'agent si vide)   "
        "🟢 Vert = Calculé / Projeté par l'agent"
    )
    leg.font = Font(name="Arial", size=8, italic=True, color="FF404040")
    leg.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row[0]].height = 22

    ws.freeze_panes = "C4"


def _build_notice_sheet(wb):
    ws = wb.create_sheet("Notice & Instructions")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 90

    lines = [
        ("NOTICE D'UTILISATION DU TEMPLATE", "title"),
        ("", None),
        ("Ce fichier est le template officiel pour l'évaluation d'entreprise par DCF et Multiples.", "body"),
        ("", None),
        ("RÈGLES DE SAISIE", "section"),
        ("• Saisissez uniquement dans les cellules bleues (historique N-3 à N-1).", "body"),
        ("• Les cellules jaunes sont optionnelles : l'agent les estimera si vides.", "body"),
        ("• Les cellules vertes sont projetées automatiquement par l'agent.", "body"),
        ("", None),
        ("UNITÉS", "section"),
        ("• Tous les montants en kMAD (milliers de dirhams) sauf indication contraire.", "body"),
        ("• Les taux en valeur décimale : 30% → 0.30", "body"),
        ("", None),
        ("TAUX DE CROISSANCE", "section"),
        ("• Si laissé vide, l'agent l'estime à partir des données historiques et du secteur.", "body"),
        ("• Vous pouvez fournir un taux court terme (N+1..N+5) et un taux terminal distincts.", "body"),
        ("", None),
        ("COMPARABLES SECTORIELS", "section"),
        ("• L'agent recherchera 1 à 3 entreprises comparables cotées.", "body"),
        ("• Il proposera leurs multiples EV/EBITDA, P/E et EV/CA pour validation.", "body"),
    ]

    ws.row_dimensions[1].height = 10
    for i, (text, style) in enumerate(lines, start=2):
        c = ws.cell(row=i, column=1, value=text)
        if style == "title":
            c.font = Font(name="Arial", bold=True, size=14, color=C_HEADER_FG)
            c.fill = _header_fill()
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[i].height = 28
        elif style == "section":
            c.font = Font(name="Arial", bold=True, size=10, color="FF1F3864")
            c.fill = _section_fill()
            ws.row_dimensions[i].height = 18
        elif style == "body":
            c.font = Font(name="Arial", size=9)
            ws.row_dimensions[i].height = 16
        if style:
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def generer_template(output_path: str = "template_evaluation.xlsx") -> str:
    """Génère le template Excel et retourne son chemin absolu."""
    wb = Workbook()
    _build_financial_sheet(wb)
    _build_notice_sheet(wb)
    wb.save(output_path)
    return os.path.abspath(output_path)